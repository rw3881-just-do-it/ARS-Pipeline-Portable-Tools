#
# Name: excel2ars.py
#
# Description: Reads information in an Excel file (based on "ARS Template.xlsx")
#              and creates an output file containing an ARS ReportingEvent. The
#              output file will be:
#              - created in the same folder as the original Excel file,
#              - named according to the name of the input Excel file, and
#              - in either JSON or YAML format, depending on the specified
#                output format (default is JSON).
#              For example, an Excel file called "Study 1 CSR.xlsx" will generate an
#              output file called either "Study 1 CSR.json" (with -of unspecified or
#              specified as json) or "Study 1 CSR.yaml" (with -of specified as yaml).
#              The id, version and name of the ReportingEvent will be as specified 
#              on the ReportEvent sheet of the Excel file.
#
# Usage: python excel2ars.py -x <path to Excel file> [-of <output format: json|yaml>]
#
# Examples: 
#   > python excel2ars.py -x '..\..\workfiles\examples\Sprint 12 Examples.xlsx'
#   Creates ..\..\workfiles\examples\Sprint 12 Examples.json
#
#   > python excel2ars.py -x '..\..\workfiles\examples\Sprint 12 Examples.xlsx' -of yaml
#   Creates ..\..\workfiles\examples\Sprint 12 Examples.yaml
#

import os
import sys
import argparse
import re
from linkml_runtime.dumpers import yaml_dumper, json_dumper
from openpyxl import load_workbook, Workbook

# Add the directory containing this script to path so that the ars_ldm
# module (from the CDISC project) can be found.
import os
lib_path = os.path.abspath(os.path.dirname(__file__))
if lib_path not in sys.path:
    sys.path.append(lib_path)

from ars_ldm import *


def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-x", "--excel_file", help="Excel file containing details of reporting event"
    )
    parser.add_argument(
        "-of", "--output_format", help="[Optional] Output format for the generated file (json or yaml). Default is json.", default="json"
    )
    args = parser.parse_args()
    return args


args = parse_arguments()

if not args.excel_file:
    print("The name of the Excel file must be specified.")
    exit(1)

if not args.output_format in ["yaml", "json"]:
    print("The output format must be specified as 'json' or 'yaml'.")
    exit(1)

filename = os.path.split(args.excel_file)[-1]

if "".join(filename.split(".")[-1]).lower() != "xlsx":
    print("Only xlsx files are valid.")
    exit(1)

rptevtfname = "".join(filename.split(".")[0])

wb: Workbook
wb = load_workbook(filename=args.excel_file)


def get_nlist(sheetname: str, clist: str, vmap: dict, values: list[list]) -> NestedList:
    nlists = {}
    lvl = 0
    for value in values:
        if value[vmap["listItem_order"]] == 1:
            nlists[value[vmap["listItem_level"]]] = NestedList(
                listItems=[
                    OrderedListItem(
                        level=value[vmap["listItem_level"]],
                        name=value[vmap["listItem_name"]],
                        description=value[vmap["listItem_description"]]
                        if "listItem_description" in vmap
                        else None,
                        label=value[vmap["listItem_label"]]
                        if "listItem_label" in vmap
                        else None,
                        order=value[vmap["listItem_order"]],
                        analysisId=value[vmap["listItem_analysisId"]],
                        outputId=value[vmap["listItem_outputId"]],
                    )
                ]
            )
            lvl = value[vmap["listItem_level"]]
        elif value[vmap["listItem_level"]] <= lvl:
            if value[vmap["listItem_level"]] < lvl:
                for _i in range(lvl)[: (value[vmap["listItem_level"]] - 1) : -1]:
                    nlists[_i].listItems[-1].sublist = nlists.pop(_i + 1)
                    lvl = value[vmap["listItem_level"]]
            nlists[value[vmap["listItem_level"]]].listItems.append(
                OrderedListItem(
                    level=value[vmap["listItem_level"]],
                    name=value[vmap["listItem_name"]],
                    description=value[vmap["listItem_description"]]
                    if "listItem_description" in vmap
                    else None,
                    label=value[vmap["listItem_label"]]
                    if "listItem_label" in vmap
                    else None,
                    order=value[vmap["listItem_order"]],
                    analysisId=value[vmap["listItem_analysisId"]],
                    outputId=value[vmap["listItem_outputId"]],
                )
            )
        else:
            print(
                f"Unexpected entry in {sheetname} content list '{clist}': level={value[vmap['listItem_level']]}, order={value[vmap['listItem_order']]}"
            )
    else:
        for _i in range(lvl)[:0:-1]:
            nlists[_i].listItems[-1].sublist = nlists.pop(_i + 1)

    return nlists[1]


clists = {}
for sheetname in ["MainListOfContents", "OtherListsOfContents"]:
    wsLopx = wb[sheetname]
    mapLopx = {col.value: col.column - 1 for col in tuple(wsLopx.rows)[0]}
    lcols = len([1 for k in mapLopx.keys() if not k.startswith("listItem_")])
    for value in wsLopx.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue
        if value[mapLopx["name"]] in clists:
            clists[value[mapLopx["name"]]]["values"].append(
                value[mapLopx["listItem_level"] :]
            )
            if (
                "description" in mapLopx
                and value[mapLopx["description"]]
                != clists[value[mapLopx["name"]]]["description"]
            ):
                print(
                    f"Inconsistent description found for {value[mapLopx['name']]} "
                    + f"content list in {sheetname} sheet: '{value[mapLopx['description']]}'"
                    + f"/'{clists[value[mapLopx['name']]]['description']}'"
                )
            if (
                "label" in mapLopx
                and value[mapLopx["label"]] != clists[value[mapLopx["name"]]]["label"]
            ):
                print(
                    f"Inconsistent label found for {value[mapLopx['name']]} "
                    + f"content list in {sheetname} sheet: '{value[mapLopx['label']]}'"
                    + f"/'{clists[value[mapLopx['name']]]['label']}'"
                )
        else:
            clists[value[mapLopx["name"]]] = {
                "sheetname": sheetname,
                "main": (sheetname == "MainListOfContents"),
                "description": value[mapLopx["description"]]
                if "description" in mapLopx
                else None,
                "label": value[mapLopx["label"]] if "label" in mapLopx else None,
                "vmap": {
                    k: v - lcols
                    for k, v in mapLopx.items()
                    if k.startswith("listItem_")
                },
                "values": [value[mapLopx["listItem_level"] :]],
            }

found_main: bool = False

for k, v in clists.items():
    if clists[k]["main"]:
        if found_main:
            print(
                f"Only one main list of contents allowed. Content list '{k}' ignored."
            )
            x = clists[k].pop("main")
            continue
        else:
            found_main = True
    clists[k]["clist"] = ListOfContents(
        name=k,
        description=v["description"] if "description" in v else None,
        label=v["label"] if "label" in v else None,
        contentsList=get_nlist(
            sheetname=v["sheetname"], clist=k, vmap=v["vmap"], values=v["values"]
        ),
    )


def get_content_lists(main: bool) -> list[ListOfContents]:
    return [v["clist"] for v in clists.values() if "main" in v and v["main"] == main]


wsRptEvt = wb["ReportingEvent"]
mapRptEvt = {col.value: col.column - 1 for col in tuple(wsRptEvt.rows)[0]}
for value in wsRptEvt.iter_rows(min_row=2, max_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    rptevt = ReportingEvent(
        id=value[mapRptEvt["id"]],
        version=value[mapRptEvt["version"]] if "version" in mapRptEvt else None,
        name=value[mapRptEvt["name"]],
        description=value[mapRptEvt["description"]]
        if "description" in mapRptEvt
        else None,
        label=value[mapRptEvt["label"]] if "label" in mapRptEvt else None,
        mainListOfContents=get_content_lists(main=True)[0],
        otherListsOfContents=get_content_lists(main=False),
    )


wsRefDoc = wb["ReferenceDocuments"]
mapRefDoc = {col.value: col.column - 1 for col in tuple(wsRefDoc.rows)[0]}

for value in wsRefDoc.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    refdoc = ReferenceDocument(
        id=value[mapRefDoc["id"]],
        name=value[mapRefDoc["name"]],
        description=value[mapRefDoc["description"]]
        if "description" in mapRefDoc
        else None,
        label=value[mapRefDoc["label"]] if "label" in mapRefDoc else None,
        location=value[mapRefDoc["location"]] if "location" in mapRefDoc else None,
    )
    rptevt.referenceDocuments.append(refdoc)

wsCat = wb["Categorizations"]
mapCat = {col.value: col.column - 1 for col in tuple(wsCat.rows)[0]}
catnid = ""
cats = {}
catns = {}

for value in wsCat.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    if value[mapCat["id"]] != catnid:
        if value[mapCat["parent_category_id"]] is None:
            rptevt.analysisOutputCategorizations.append(
                AnalysisOutputCategorization(
                    id=value[mapCat["id"]],
                    label=value[mapCat["label"]] if "label" in mapCat else None,
                    categories=[
                        AnalysisOutputCategory(
                            id=value[mapCat["category_id"]],
                            label=value[mapCat["category_label"]],
                        )
                    ],
                )
            )
            catns[value[mapCat["id"]]] = rptevt.analysisOutputCategorizations[-1]
        else:
            if value[mapCat["parent_category_id"]] in cats:
                cats[value[mapCat["parent_category_id"]]].subCategorizations.append(
                    AnalysisOutputCategorization(
                        id=value[mapCat["id"]],
                        label=value[mapCat["label"]] if "label" in mapCat else None,
                        categories=[
                            AnalysisOutputCategory(
                                id=value[mapCat["category_id"]],
                                label=value[mapCat["category_label"]],
                            )
                        ],
                    )
                )
                catns[value[mapCat["id"]]] = cats[
                    value[mapCat["parent_category_id"]]
                ].subCategorizations[-1]
            else:
                print(
                    f"Unknown parent_category_id: {value[mapCat['parent_category_id']]}"
                )

        catnid = value[mapCat["id"]]
        pcatid = value[mapCat["parent_category_id"]]

    else:

        catns[value[mapCat["id"]]].categories.append(
            AnalysisOutputCategory(
                id=value[mapCat["category_id"]], label=value[mapCat["category_label"]]
            )
        )
        if value[mapCat["parent_category_id"]] != pcatid:
            print(
                f"Inconsistent parent_category_id values defined for categorization {value[mapCat['id']]}: {pcatid} / {value[mapCat['parent_category_id']]}"
            )

    cats[value[mapCat["category_id"]]] = catns[value[mapCat["id"]]].categories[-1]

wsGDS = wb["GlobalDisplaySections"]
mapGDS = {col.value: col.column - 1 for col in tuple(wsGDS.rows)[0]}
sidx = {}
dsubsects = {}
for value in wsGDS.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    ssect = DisplaySubSection(
        id=value[mapGDS["subSection_id"]], text=value[mapGDS["subSection_text"]]
    )

    if value[mapGDS["sectionType"]] in sidx:
        sidx[value[mapGDS["sectionType"]]].subSections.append(ssect)
    else:
        rptevt.globalDisplaySections.append(
            GlobalDisplaySection(
                sectionType=value[mapGDS["sectionType"]], subSections=[ssect]
            )
        )
        sidx[value[mapGDS["sectionType"]]] = rptevt.globalDisplaySections[-1]

    dsubsects[value[mapGDS["subSection_id"]]] = sidx[
        value[mapGDS["sectionType"]]
    ].subSections[-1]

wsTermEx = wb["TerminologyExtensions"]
mapTermEx = {col.value: col.column - 1 for col in tuple(wsTermEx.rows)[0]}
termexidx = {}
spterms = {}
for value in wsTermEx.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    spterm = SponsorTerm(
        id=value[mapTermEx["sponsorTerm_id"]],
        submissionValue=value[mapTermEx["sponsorTerm_submissionValue"]],
        description=value[mapTermEx["sponsorTerm_description"]],
    )

    if value[mapTermEx["id"]] in termexidx:
        termexidx[value[mapTermEx["id"]]].sponsorTerms.append(spterm)
        if (
            value[mapTermEx["enumeration"]]
            != termexidx[value[mapTermEx["id"]]].enumeration
        ):
            print(
                f"More than one enumeration specified for terminology extension {value[mapTermEx['id']]}: {termexidx[value[mapTermEx['id']]].enumeration} / {value[mapTermEx['enumeration']]}"
            )
    else:
        rptevt.terminologyExtensions.append(
            TerminologyExtension(
                id=value[mapTermEx["id"]],
                enumeration=value[mapTermEx["enumeration"]],
                sponsorTerms=[spterm],
            )
        )
        termexidx[value[mapTermEx["id"]]] = rptevt.terminologyExtensions[-1]

    spterms[value[mapTermEx["enumeration"]]] = termexidx[
        value[mapTermEx["id"]]
    ].sponsorTerms[-1]

wsAnSet = wb["AnalysisSets"]
mapAnSet = {col.value: col.column - 1 for col in tuple(wsAnSet.rows)[0]}

asid = ""
anset: AnalysisSet = None
wcs = {}

for value in wsAnSet.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[mapAnSet["compoundExpression_logicalOperator"]] is None:
        if value[mapAnSet["condition_variable"]] is not None:
            wc = WhereClause(
                level=value[mapAnSet["level"]],
                order=value[mapAnSet["order"]],
                condition=WhereClauseCondition(
                    dataset=value[mapAnSet["condition_dataset"]],
                    variable=value[mapAnSet["condition_variable"]],
                    comparator=value[mapAnSet["condition_comparator"]],
                    value=str(value[mapAnSet["condition_value"]]).split(" | ")
                    if value[mapAnSet["condition_value"]] is not None
                    and str(value[mapAnSet["condition_value"]]).find(" | ") > -1
                    else value[mapAnSet["condition_value"]],
                )
                if value[mapAnSet["condition_variable"]] is not None
                else None,
            )
        elif (
            "compoundExpression_subClauseId" in mapAnSet
            and value[mapAnSet["compoundExpression_subClauseId"]] is not None
        ):
            wc = ReferencedAnalysisSet(
                level=value[mapAnSet["level"]],
                order=value[mapAnSet["order"]],
                subClauseId=value[mapAnSet["compoundExpression_subClauseId"]],
            )
        else:
            print(
                f"Both condition variable and subClauseId are missing for an analysis set {value[mapAnSet['id']]}: level={value[mapAnSet['level']]}, order={value[mapAnSet['order']]}"
            )
    else:
        wc = WhereClause(
            level=value[mapAnSet["level"]],
            order=value[mapAnSet["order"]],
            compoundExpression=CompoundSetExpression(
                logicalOperator=value[mapAnSet["compoundExpression_logicalOperator"]],
                whereClauses=[],
            )
            if value[mapAnSet["compoundExpression_logicalOperator"]] is not None
            else None,
        )
        wcs[value[mapAnSet["level"]]] = wc

    if value[mapAnSet["id"]] != asid:
        if asid != "":
            rptevt.analysisSets.append(anset)
        anset = AnalysisSet(
            id=value[mapAnSet["id"]],
            name=value[mapAnSet["name"]],
            description=value[mapAnSet["description"]]
            if "description" in mapAnSet
            else None,
            label=value[mapAnSet["label"]] if "label" in mapAnSet else None,
            level=wc.level,
            order=wc.order,
            condition=wc.condition,
            compoundExpression=wc.compoundExpression,
        )
        asid = value[mapAnSet["id"]]
    else:
        if isinstance(
            wcs[value[mapAnSet["level"]] - 1].compoundExpression.whereClauses, list
        ):
            wcs[value[mapAnSet["level"]] - 1].compoundExpression.whereClauses.extend(
                [wc]
            )
        else:
            wcs[value[mapAnSet["level"]] - 1].compoundExpression.whereClauses = [wc]
else:
    rptevt.analysisSets.append(anset)

wsAnGrp = wb["AnalysisGroupings"]
mapAnGrp = {col.value: col.column - 1 for col in tuple(wsAnGrp.rows)[0]}
grpngid = ""
grpid = ""

grpng = GroupingFactor(id="dummy", dataDriven=False, name="dummy")
for value in wsAnGrp.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if str(value[mapAnGrp["id"]]) != grpngid:
        wcs = {}

    if value[mapAnGrp["group_id"]] is not None:
        if value[mapAnGrp["group_compoundExpression_logicalOperator"]] is None:
            if value[mapAnGrp["group_condition_variable"]] is not None:
                wc = WhereClause(
                    level=value[mapAnGrp["group_level"]],
                    order=value[mapAnGrp["group_order"]],
                    condition=WhereClauseCondition(
                        dataset=value[mapAnGrp["group_condition_dataset"]],
                        variable=value[mapAnGrp["group_condition_variable"]],
                        comparator=value[mapAnGrp["group_condition_comparator"]],
                        value=str(value[mapAnGrp["group_condition_value"]]).split(" | ")
                        if value[mapAnGrp["group_condition_value"]] is not None
                        and str(value[mapAnGrp["group_condition_value"]]).find(" | ")
                        > -1
                        else value[mapAnGrp["group_condition_value"]],
                    )
                    if value[mapAnGrp["group_condition_variable"]] is not None
                    else None,
                )
            elif (
                "group_compoundExpression_subClauseId" in mapAnGrp
                and value[mapAnGrp["group_compoundExpression_subClauseId"]] is not None
            ):
                wc = ReferencedGroup(
                    level=value[mapAnGrp["group_level"]],
                    order=value[mapAnGrp["group_order"]],
                    subClauseId=value[mapAnGrp["group_compoundExpression_subClauseId"]],
                )
            else:
                print(
                    f"Both condition variable and subClauseId are missing for a group {value[mapAnGrp['group_id']]}: level={value[mapAnGrp['group_level']]}, order={value[mapAnGrp['group_order']]}"
                )
        else:
            wc = WhereClause(
                level=value[mapAnGrp["group_level"]],
                order=value[mapAnGrp["group_order"]],
                compoundExpression=CompoundGroupExpression(
                    logicalOperator=value[
                        mapAnGrp["group_compoundExpression_logicalOperator"]
                    ],
                    whereClauses=[],
                )
                if value[mapAnGrp["group_compoundExpression_logicalOperator"]]
                is not None
                else None,
            )
            wcs[value[mapAnGrp["group_level"]]] = wc

        if value[mapAnGrp["group_id"]] != grpid:
            if grpid != "":
                grpng.groups.append(grp)
            grp = Group(
                id=value[mapAnGrp["group_id"]],
                name=value[mapAnGrp["group_name"]],
                description=value[mapAnGrp["group_description"]]
                if "group_description" in mapAnGrp
                else None,
                label=value[mapAnGrp["group_label"]]
                if "group_label" in mapAnGrp
                else None,
                level=wc.level,
                order=wc.order,
                condition=wc.condition,
                compoundExpression=wc.compoundExpression,
            )
            grpid = value[mapAnGrp["group_id"]]
        else:
            if isinstance(
                wcs[value[mapAnGrp["group_level"]] - 1].compoundExpression.whereClauses,
                list,
            ):
                wcs[
                    value[mapAnGrp["group_level"]] - 1
                ].compoundExpression.whereClauses.extend([wc])
            else:
                wcs[
                    value[mapAnGrp["group_level"]] - 1
                ].compoundExpression.whereClauses = [wc]
    else:
        if grpid != "":
            grpng.groups.append(grp)
        wc = None
        grpid = ""

    # If this is a new grouping...
    if str(value[mapAnGrp["id"]]) != grpngid:
        # Store what's been built for the previous grouping
        if grpng.id != "dummy":
            rptevt.analysisGroupings.append(grpng)
        grpngid = str(value[mapAnGrp["id"]])
        # Then process the new grouping
        grpng = GroupingFactor(
            id=grpngid,
            name=value[mapAnGrp["name"]],
            description=value[mapAnGrp["description"]]
            if "description" in mapAnGrp
            else None,
            label=value[mapAnGrp["label"]] if "label" in mapAnGrp else None,
            groupingDataset=value[mapAnGrp["groupingDataset"]],
            groupingVariable=value[mapAnGrp["groupingVariable"]],
            dataDriven=value[mapAnGrp["dataDriven"]],
            groups=[],
        )
#    else:
#        grpng.groups.append(DataGroup(id=value[mapAnGrp["group_id"]],label=value[mapAnGrp["group_label"]],level=value[mapAnGrp["group_level"]],order=value[mapAnGrp["group_order"]],condition=WhereClauseCondition(dataset=value[mapAnGrp["group_condition_dataset"]],variable=value[mapAnGrp["group_condition_variable"]],comparator=value[mapAnGrp["group_condition_comparator"]],value=str(value[mapAnGrp["group_condition_value"]]).split(" | ") if value[mapAnGrp["group_condition_value"]] is not None and str(value[mapAnGrp["group_condition_value"]]).find(" | ") > -1 else value[mapAnGrp["group_condition_value"]])))
else:
    if grpid != "":
        grpng.groups.append(grp)
    rptevt.analysisGroupings.append(grpng)

wsDss = wb["DataSubsets"]
mapDss = {col.value: col.column - 1 for col in tuple(wsDss.rows)[0]}
dss: DataSubset = None
dssid = ""
wcs = {}

for value in wsDss.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[mapDss["compoundExpression_logicalOperator"]] is None:
        if value[mapDss["condition_variable"]] is not None:
            wc = WhereClause(
                level=value[mapDss["level"]],
                order=value[mapDss["order"]],
                condition=WhereClauseCondition(
                    dataset=value[mapDss["condition_dataset"]],
                    variable=value[mapDss["condition_variable"]],
                    comparator=value[mapDss["condition_comparator"]],
                    value=str(value[mapDss["condition_value"]]).split(" | ")
                    if value[mapDss["condition_value"]] is not None
                    and str(value[mapDss["condition_value"]]).find(" | ") > -1
                    else value[mapDss["condition_value"]],
                )
                if value[mapDss["condition_variable"]] is not None
                else None,
            )
        elif (
            "compoundExpression_subClauseId" in mapDss
            and value[mapDss["compoundExpression_subClauseId"]] is not None
        ):
            wc = ReferencedDataSubset(
                level=value[mapDss["level"]],
                order=value[mapDss["order"]],
                subClauseId=value[mapDss["compoundExpression_subClauseId"]],
            )
        else:
            print(
                f"Both condition variable and subClauseId are missing for data subset {value[mapDss['id']]}: level={value[mapDss['level']]}, order={value[mapDss['order']]}"
            )
    else:
        wc = WhereClause(
            level=value[mapDss["level"]],
            order=value[mapDss["order"]],
            compoundExpression=CompoundSubsetExpression(
                logicalOperator=value[mapDss["compoundExpression_logicalOperator"]],
                whereClauses=[],
            )
            if value[mapDss["compoundExpression_logicalOperator"]] is not None
            else None,
        )
        wcs[value[mapDss["level"]]] = wc

    if value[mapDss["id"]] != dssid:
        if dssid != "":
            rptevt.dataSubsets.append(dss)
        dss = DataSubset(
            id=value[mapDss["id"]],
            name=value[mapDss["name"]],
            description=value[mapDss["description"]]
            if "description" in mapDss
            else None,
            label=value[mapDss["label"]],
            level=wc.level,
            order=wc.order,
            condition=wc.condition,
            compoundExpression=wc.compoundExpression,
        )
        dssid = value[mapDss["id"]]
    else:
        if isinstance(
            wcs[value[mapDss["level"]] - 1].compoundExpression.whereClauses, list
        ):
            wcs[value[mapDss["level"]] - 1].compoundExpression.whereClauses.extend([wc])
        else:
            wcs[value[mapDss["level"]] - 1].compoundExpression.whereClauses = [wc]
else:
    if dss:
        rptevt.dataSubsets.append(dss)

def read_operation_result(value: list, mapRslts: dict) -> OperationResult:
    return OperationResult(
        operationId=value[mapRslts["operation_id"]],
        rawValue=value[mapRslts["rawValue"]],
        formattedValue=value[mapRslts["formattedValue"]],
        resultGroups=[
            ResultGroup(
                groupingId=value[x], groupId=value[y], groupValue=value[z]
            )
            for x, y, z in [
                [
                    mapRslts[rg + "_groupingId"],
                    mapRslts[rg + "_groupId"],
                    mapRslts[rg + "_groupValue"],
                ]
                for rg in sorted(set(
                    k.split("_")[0]
                    for k in mapRslts.keys()
                    if k.startswith("resultGroup")
                ))
            ]
            if value[x] is not None
        ],
    )

results = {}

if "AnalysisResults" in wb.sheetnames:
    wsRslts = wb["AnalysisResults"]
    mapRslts = {col.value: col.column - 1 for col in tuple(wsRslts.rows)[0]}
    anid = ""
    for value in wsRslts.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue
        if str(value[mapRslts["id"]]) != anid:
            results[value[mapRslts["id"]]] = [
                read_operation_result(value, mapRslts)
            ]
            anid = str(value[mapRslts["id"]])
        else:
            results[value[mapRslts["id"]]].append(
                read_operation_result(value, mapRslts)
            )

def get_docrefs(sheetname: str) -> dict:

    wsDocRef = wb[sheetname]
    mapDocRef = {col.value: col.column - 1 for col in tuple(wsDocRef.rows)[0]}

    docrefs = {
        "sheetname": sheetname,
        "Documentation": dict(),
        "ProgrammingCode": dict(),
    }

    docidx = {
        "sheetname": sheetname,
        "Documentation": dict(),
        "ProgrammingCode": dict(),
    }

    for value in wsDocRef.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        pageref = None

        if value[mapDocRef["pageRef_refType"]] is not None:
            if value[mapDocRef["pageRef_refType"]] == "NamedDestination":
                pageref = PageNameRef(
                    refType=value[mapDocRef["pageRef_refType"]],
                    label=value[mapDocRef["pageRef_label"]],
                    pageNames=value[mapDocRef["pageRef_pages"]].split("|")
                    if value[mapDocRef["pageRef_pages"]] is not None
                    and str(value[mapDocRef["pageRef_pages"]]).find("|") > -1
                    else str(value[mapDocRef["pageRef_pages"]])
                    if value[mapDocRef["pageRef_pages"]] is not None
                    else None,
                )
            elif value[mapDocRef["pageRef_refType"]] == "PhysicalRef":
                if re.search("^[0-9]+-[0-9]+$", str(value[mapDocRef["pageRef_pages"]])):
                    pageref = PageNumberRangeRef(
                        refType=value[mapDocRef["pageRef_refType"]],
                        label=value[mapDocRef["pageRef_label"]],
                        firstPage=value[mapDocRef["pageRef_pages"]].split("-")[0],
                        lastPage=value[mapDocRef["pageRef_pages"]].split("-")[1],
                    )
                elif re.search(
                    "^[0-9]+(\|[0-9]+)*$", str(value[mapDocRef["pageRef_pages"]])
                ):
                    pageref = PageNumberListRef(
                        refType=value[mapDocRef["pageRef_refType"]],
                        label=value[mapDocRef["pageRef_label"]],
                        pageNumbers=value[mapDocRef["pageRef_pages"]].split("|")
                        if value[mapDocRef["pageRef_pages"]] is not None
                        and str(value[mapDocRef["pageRef_pages"]]).find("|") > -1
                        else str(value[mapDocRef["pageRef_pages"]])
                        if value[mapDocRef["pageRef_pages"]] is not None
                        else None,
                    )
                else:
                    print(
                        f"Invalid pageRef_pages value on {sheetname} sheet: {value[mapDocRef['pageRef_pages']]}"
                    )
            else:
                print(
                    f"Invalid pageRef_refType value on {sheetname} sheet: {value[mapDocRef['pageRef_refType']]}"
                )

        idtype = (
            "method_id"
            if sheetname == "AnalysisMethodDocumentRefs"
            else sheetname.replace("DocumentRefs", "").lower() + "_id"
        )

        if value[mapDocRef[idtype]] in docrefs[value[mapDocRef["referenceType"]]]:
            if (
                value[mapDocRef["refDocumentId"]]
                not in docidx[value[mapDocRef["referenceType"]]][
                    value[mapDocRef[idtype]]
                ]
            ):
                docrefs[value[mapDocRef["referenceType"]]][
                    value[mapDocRef[idtype]]
                ].append(
                    DocumentReference(
                        referenceDocumentId=value[mapDocRef["refDocumentId"]]
                    )
                )
                docidx[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]][
                    value[mapDocRef["refDocumentId"]]
                ] = docrefs[value[mapDocRef["referenceType"]]][
                    value[mapDocRef[idtype]]
                ][
                    -1
                ]
                docidx[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]][
                    value[mapDocRef["refDocumentId"]]
                ].pageRefs = ([pageref] if pageref else None)
            elif pageref:
                docidx[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]][
                    value[mapDocRef["refDocumentId"]]
                ].pageRefs.append(pageref)
        else:
            docrefs[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]] = [
                DocumentReference(
                    referenceDocumentId=value[mapDocRef["refDocumentId"]]
                )
            ]
            docidx[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]] = {
                value[mapDocRef["refDocumentId"]]: docrefs[
                    value[mapDocRef["referenceType"]]
                ][value[mapDocRef[idtype]]][-1]
            }
            docidx[value[mapDocRef["referenceType"]]][value[mapDocRef[idtype]]][
                value[mapDocRef["refDocumentId"]]
            ].pageRefs = ([pageref] if pageref else None)

    return docrefs


def get_params(sheetname: str, template: bool = False) -> dict:

    wsParams = wb[sheetname]
    mapParams = {col.value: col.column - 1 for col in tuple(wsParams.rows)[0]}

    params = {}

    for value in wsParams.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        if template:
            param = TemplateCodeParameter(
                name=value[mapParams["parameter_name"]],
                description=value[mapParams["parameter_description"]]
                if "parameter_description" in mapParams
                else None,
                label=value[mapParams["parameter_label"]]
                if "parameter_label" in mapParams
                else None,
                valueSource=value[mapParams["parameter_valueSource"]],
                value=value[mapParams["parameter_value"]].split("|")
                if value[mapParams["parameter_value"]] is not None
                and str(value[mapParams["parameter_value"]]).find("|") > -1
                else str(value[mapParams["parameter_value"]])
                if value[mapParams["parameter_value"]]
                else None,
            )
        else:
            param = AnalysisOutputCodeParameter(
                name=value[mapParams["parameter_name"]],
                description=value[mapParams["parameter_description"]]
                if "parameter_description" in mapParams
                else None,
                label=value[mapParams["parameter_label"]]
                if "parameter_label" in mapParams
                else None,
                value=value[mapParams["parameter_value"]],
            )

        idtype = (
            "method_id"
            if sheetname == "AnalysisMethodCodeParameters"
            else sheetname.replace("CodeParameters", "").lower() + "_id"
        )

        if value[mapParams[idtype]] in params:
            params[value[mapParams[idtype]]].append(param)
        else:
            params[value[mapParams[idtype]]] = [param]

    return params


def get_progcode(
    sheetname: str, docrefs: dict, params: dict, template: bool = False
) -> dict:

    wsProgCode = wb[sheetname]
    mapProgCode = {col.value: col.column - 1 for col in tuple(wsProgCode.rows)[0]}

    progcode = {}

    for value in wsProgCode.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        idtype = (
            "method_id"
            if template
            else sheetname.replace("ProgrammingCode", "").lower() + "_id"
        )

        if value[mapProgCode["specifiedAs"]] == "Code":
            if template:
                progcode[value[mapProgCode[idtype]]] = AnalysisProgrammingCodeTemplate(
                    context=value[mapProgCode["context"]],
                    code=value[mapProgCode["templateCode"]],
                    parameters=params[value[mapProgCode[idtype]]]
                    if value[mapProgCode[idtype]] in params
                    else None,
                )
            else:
                progcode[value[mapProgCode[idtype]]] = AnalysisOutputProgrammingCode(
                    context=value[mapProgCode["context"]],
                    code=value[mapProgCode["code"]],
                    parameters=params[value[mapProgCode[idtype]]]
                    if value[mapProgCode[idtype]] in params
                    else None,
                )
        elif value[mapProgCode["specifiedAs"]] == "DocumentRef":
            if value[mapProgCode[idtype]] in docrefs["ProgrammingCode"]:
                if template:
                    progcode[
                        value[mapProgCode[idtype]]
                    ] = AnalysisProgrammingCodeTemplate(
                        context=value[mapProgCode["context"]],
                        documentRef=docrefs["ProgrammingCode"][
                            value[mapProgCode[idtype]]
                        ][0],
                        parameters=params[value[mapProgCode[idtype]]]
                        if value[mapProgCode[idtype]] in params
                        else None,
                    )
                else:
                    progcode[
                        value[mapProgCode[idtype]]
                    ] = AnalysisOutputProgrammingCode(
                        context=value[mapProgCode["context"]],
                        documentRef=docrefs["ProgrammingCode"][
                            value[mapProgCode[idtype]]
                        ][0],
                        parameters=params[value[mapProgCode[idtype]]]
                        if value[mapProgCode[idtype]] in params
                        else None,
                    )
            else:
                print(
                    f"Programming code specified as DocumentRef for {value[mapProgCode[idtype]]} on {sheetname} sheet, but no matching ProgrammingCode DocumentRef was found on the {docrefs['sheetname']} sheet."
                )
        else:
            print(
                f"Invalid specifiedAs value on the {sheetname} sheet: {value[mapProgCode['specifiedAs']]}"
            )

    return progcode


mparams = get_params(sheetname="AnalysisMethodCodeParameters", template=True)
mdocrefs = get_docrefs(sheetname="AnalysisMethodDocumentRefs")
mprogcode = get_progcode(
    sheetname="AnalysisMethodCodeTemplate",
    docrefs=mdocrefs,
    params=mparams,
    template=True,
)

wsMth = wb["AnalysisMethods"]
mapMth = {col.value: col.column - 1 for col in tuple(wsMth.rows)[0]}
mthid = ""
method = AnalysisMethod(
    id="dummy", name="dummy", operations=[Operation(id="dummy", name="dummy", order=0)]
)
for value in wsMth.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    # If this is a new method...
    if str(value[mapMth["id"]]) != mthid:
        # Store what's been built for the previous method
        if method.id != "dummy":
            rptevt.methods.append(method)
        mthid = str(value[mapMth["id"]])
        # Then process the new method
        method = AnalysisMethod(
            id=mthid,
            name=value[mapMth["name"]],
            description=value[mapMth["description"]]
            if "description" in mapMth
            else None,
            label=value[mapMth["label"]] if "label" in mapMth else None,
            operations=[
                Operation(
                    id=value[mapMth["operation_id"]],
                    name=value[mapMth["operation_name"]],
                    description=value[mapMth["operation_description"]]
                    if "operation_description" in mapMth
                    else None,
                    label=value[mapMth["operation_label"]]
                    if "operation_label" in mapMth
                    else None,
                    order=value[mapMth["operation_order"]],
                    resultPattern=value[mapMth["operation_resultPattern"]],
                    referencedOperationRelationships=[
                        ReferencedOperationRelationship(
                            id=value[mapMth[orrr + "_id"]],
                            referencedOperationRole=OperationRole(
                                controlledTerm=value[
                                    mapMth[orrr + "_referencedOperationRole"]
                                ]
                            )
                            if value[mapMth[orrr + "_referencedOperationRole"]]
                            in OperationRoleEnum
                            else SponsorOperationRole(
                                sponsorTermId=value[
                                    mapMth[orrr + "_referencedOperationRole"]
                                ]
                            ),
                            operationId=value[mapMth[orrr + "_operationId"]],
                            analysisId=value[mapMth[orrr + "_analysisId"]]
                            if orrr + "_analysisId" in mapMth
                            else None,
                            description=value[mapMth[orrr + "_description"]]
                            if orrr + "_description" in mapMth
                            else None,
                        )
                        for orrr in sorted(
                            set(
                                [
                                    "_".join(k.split("_")[0:2])
                                    for k in mapMth.keys()
                                    if k.startswith(
                                        "operation_referencedResultRelationships"
                                    )
                                ]
                            )
                        )
                    ]
                    if "operation_referencedResultRelationships1_id" in mapMth
                    and value[mapMth["operation_referencedResultRelationships1_id"]]
                    is not None
                    else None,
                )
            ],
            documentRefs=mdocrefs["Documentation"][mthid]
            if mthid in mdocrefs["Documentation"]
            else None,
            codeTemplate=mprogcode[mthid] if mthid in mprogcode else None,
        )
    else:
        method.operations.append(
            Operation(
                id=value[mapMth["operation_id"]],
                name=value[mapMth["operation_name"]],
                description=value[mapMth["operation_description"]]
                if "operation_description" in mapMth
                else None,
                label=value[mapMth["operation_label"]]
                if "operation_label" in mapMth
                else None,
                order=value[mapMth["operation_order"]],
                resultPattern=value[mapMth["operation_resultPattern"]],
                referencedOperationRelationships=[
                    ReferencedOperationRelationship(
                        id=value[mapMth[orrr + "_id"]],
                        referencedOperationRole=OperationRole(
                            controlledTerm=value[
                                mapMth[orrr + "_referencedOperationRole"]
                            ]
                        )
                        if value[mapMth[orrr + "_referencedOperationRole"]]
                        in OperationRoleEnum
                        else SponsorOperationRole(
                            sponsorTermId=value[
                                mapMth[orrr + "_referencedOperationRole"]
                            ]
                        ),
                        operationId=value[mapMth[orrr + "_operationId"]],
                        analysisId=value[mapMth[orrr + "_analysisId"]]
                        if orrr + "_analysisId" in mapMth
                        else None,
                        description=value[mapMth[orrr + "_description"]]
                        if orrr + "_description" in mapMth
                        else None,
                    )
                    for orrr in sorted(
                        set(
                            [
                                "_".join(k.split("_")[0:2])
                                for k in mapMth.keys()
                                if k.startswith(
                                    "operation_referencedResultRelationships"
                                )
                            ]
                        )
                    )
                ]
                if "operation_referencedResultRelationships1_id" in mapMth
                and value[mapMth["operation_referencedResultRelationships1_id"]]
                is not None
                else None,
            )
        )
else:
    rptevt.methods.append(method)

aparams = get_params(sheetname="AnalysisCodeParameters")
adocrefs = get_docrefs(sheetname="AnalysisDocumentRefs")
aprogcode = get_progcode(
    sheetname="AnalysisProgrammingCode", docrefs=adocrefs, params=aparams
)

wsAn = wb["Analyses"]
mapAn = {col.value: col.column - 1 for col in tuple(wsAn.rows)[0]}

for value in wsAn.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[mapAn["reason"]] is not None:
        if value[mapAn["reason"]] in AnalysisReasonEnum:
            anreas = AnalysisReason(controlledTerm=value[mapAn["reason"]])
        else:
            anreas = SponsorAnalysisReason(sponsorTermId=value[mapAn["reason"]])
    else:
        anreas = None

    if value[mapAn["purpose"]] is not None:
        if value[mapAn["purpose"]] in AnalysisPurposeEnum:
            anpurp = AnalysisPurpose(controlledTerm=value[mapAn["purpose"]])
        else:
            anpurp = SponsorAnalysisPurpose(sponsorTermId=value[mapAn["purpose"]])
    else:
        anpurp = None

    analysis = Analysis(
        id=value[mapAn["id"]],
        version=value[mapAn["version"]],
        name=value[mapAn["name"]],
        description=value[mapAn["description"]] if "description" in mapAn else None,
        label=value[mapAn["label"]] if "label" in mapAn else None,
        reason=anreas,
        purpose=anpurp,
        documentRefs=adocrefs["Documentation"][value[mapAn["id"]]]
        if value[mapAn["id"]] in adocrefs["Documentation"]
        else None,
        analysisSetId=value[mapAn["analysisSetId"]],
        dataSubsetId=value[mapAn["dataSubsetId"]],
        dataset=value[mapAn["dataset"]],
        variable=value[mapAn["variable"]],
        methodId=value[mapAn["method_id"]],
        referencedAnalysisOperations=[
            ReferencedAnalysisOperation(
                referencedOperationRelationshipId=value[x], analysisId=value[y]
            )
            for x, y in [
                [mapAn[z], mapAn[z] + 1]
                for z in mapAn.keys()
                if z.startswith("referencedAnalysisOperations_referencedOperationId")
                and value[mapAn[z]] is not None
            ]
        ],
        programmingCode=aprogcode[value[mapAn["id"]]]
        if value[mapAn["id"]] in aprogcode
        else None,
        results=results[value[mapAn["id"]]] if value[mapAn["id"]] in results else None,
    )
    if value[mapAn["categoryIds"]] is not None:
        if value[mapAn["categoryIds"]].find(" | ") > -1:
            for catid in value[mapAn["categoryIds"]].split(" | "):
                if catid in cats:
                    analysis.categoryIds.extend([catid])
                else:
                    print(
                        f"Unknown category_id for analysis {value[mapAn['id']]}: {catid}"
                    )
        else:
            if value[mapAn["categoryIds"]] in cats:
                analysis.categoryIds.extend([value[mapAn["categoryIds"]]])
            else:
                print(
                    f"Unknown category_id for analysis {value[mapAn['id']]}: {value[mapAn['categoryIds']]}"
                )
    grpord = 0
    for _i in [gik for gik in mapAn.keys() if gik.startswith("groupingId")]:
        if value[mapAn[_i]] is None:
            break
        grpord += 1
        analysis.orderedGroupings.append(
            OrderedGroupingFactor(
                order=grpord,
                groupingId=value[mapAn[_i]],
                resultsByGroup=value[mapAn[_i] + 1],
            )
        )
    rptevt.analyses.append(analysis)

ofiles = {}

if "OutputFiles" in wb.sheetnames:
    wsOFile = wb["OutputFiles"]
    mapOFile = {col.value: col.column - 1 for col in tuple(wsOFile.rows)[0]}

    for value in wsOFile.iter_rows(min_row=2, values_only=True):
        if all([c is None for c in value]):
            continue

        fs = OutputFile(
            name=value[mapOFile["name"]],
            description=value[mapOFile["description"]]
            if "description" in mapOFile
            else None,
            label=value[mapOFile["label"]] if "label" in mapOFile else None,
            location=value[mapOFile["location"]] if "location" in mapOFile else None,
            fileType=OutputFileType(controlledTerm=value[mapOFile["fileType"]])
            if value[mapOFile["fileType"]] in OutputFileTypeEnum
            else SponsorOutputFileType(sponsorTermId=value[mapOFile["fileType"]]),
        )
        if value[mapOFile["output_id"]] in ofiles:
            ofiles[value[mapOFile["output_id"]]].append(fs)
        else:
            ofiles[value[mapOFile["output_id"]]] = [fs]

wsDisp = wb["Displays"]
mapDisp = {col.value: col.column - 1 for col in tuple(wsDisp.rows)[0]}

displays = {}
dispid = ""
for value in wsDisp.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue

    if value[mapDisp["displaySection_subSection_text"]] is None:
        ss = OrderedSubSectionRef(
            order=value[mapDisp["displaySection_orderedSubSection_order"]],
            subSectionId=value[mapDisp["displaySection_subSection_id"]],
        )
        if value[mapDisp["displaySection_subSection_id"]] not in dsubsects:
            print(
                f"Display Subsection id {value[mapDisp['displaySection_subSection_id']]} is referenced (without text), but this id has not been defined (with text)"
            )
    else:
        ss = OrderedSubSection(
            order=value[mapDisp["displaySection_orderedSubSection_order"]],
            subSection=DisplaySubSection(
                id=value[mapDisp["displaySection_subSection_id"]],
                text=value[mapDisp["displaySection_subSection_text"]],
            ),
        )
        if value[mapDisp["displaySection_subSection_id"]] in dsubsects:
            print(
                f"Duplicate display subsection id (with text defined) found: {value[mapDisp['displaySection_subSection_id']]}"
            )
        else:
            dsubsects[value[mapDisp["displaySection_subSection_id"]]] = ss

    if str(value[mapDisp["id"]]) != dispid:
        if dispid != "":
            displays[dispid] = display
        display = OutputDisplay(
            id=value[mapDisp["id"]],
            name=value[mapDisp["name"]],
            description=value[mapDisp["description"]]
            if "description" in mapDisp
            else None,
            label=value[mapDisp["label"]] if "label" in mapDisp else None,
            version=value[mapDisp["version"]],
            displayTitle=value[mapDisp["displayTitle"]],
            displaySections=[
                DisplaySection(
                    sectionType=value[mapDisp["displaySection_sectionType"]],
                    orderedSubSections=[ss],
                )
            ],
        )
        dispid = str(value[mapDisp["id"]])
        scttype = str(value[mapDisp["displaySection_sectionType"]])
    elif value[mapDisp["displaySection_sectionType"]] != scttype:
        display.displaySections.append(
            DisplaySection(
                sectionType=value[mapDisp["displaySection_sectionType"]],
                orderedSubSections=[ss],
            )
        )
        scttype = value[mapDisp["displaySection_sectionType"]]
    else:
        display.displaySections[-1].orderedSubSections.append(ss)
else:
    displays[dispid] = display

oparams = get_params(sheetname="OutputCodeParameters")
odocrefs = get_docrefs(sheetname="OutputDocumentRefs")
oprogcode = get_progcode(
    sheetname="OutputProgrammingCode", docrefs=odocrefs, params=oparams
)

wsOutput = wb["Outputs"]
mapOutput = {col.value: col.column - 1 for col in tuple(wsOutput.rows)[0]}

for value in wsOutput.iter_rows(min_row=2, values_only=True):
    if all([c is None for c in value]):
        continue
    output = Output(
        id=value[mapOutput["id"]],
        version=value[mapOutput["version"]],
        name=value[mapOutput["name"]],
        description=value[mapOutput["description"]]
        if "description" in mapOutput
        else None,
        label=value[mapOutput["label"]] if "label" in mapOutput else None,
        displays=[
            OrderedDisplay(order=d[7:], display=displays[value[mapOutput[d + "_id"]]])
            for d in [
                k.split("_")[0] for k in mapOutput.keys() if k.startswith("display")
            ]
            if value[mapOutput[d + "_id"]] is not None
        ],
        documentRefs=odocrefs["Documentation"][value[mapOutput["id"]]]
        if value[mapOutput["id"]] in odocrefs["Documentation"]
        else None,
        programmingCode=oprogcode[value[mapOutput["id"]]]
        if value[mapOutput["id"]] in oprogcode
        else None,
        fileSpecifications=ofiles[value[mapOutput["id"]]]
        if value[mapOutput["id"]] in ofiles
        else None,
    )
    if value[mapOutput["categoryIds"]] is not None:
        if value[mapOutput["categoryIds"]].find(" | ") > -1:
            for catid in value[mapOutput["categoryIds"]].split(" | "):
                if catid in cats:
                    output.categoryIds.extend([catid])
                else:
                    print(
                        f"Unknown category_id for output {value[mapOutput['categoryIds']]}: {catid}"
                    )
        else:
            if value[mapOutput["categoryIds"]] in cats:
                output.categoryIds.extend([value[mapOutput["categoryIds"]]])
            else:
                print(
                    f"Unknown category_id for output {value[mapOutput['id']]}: {value[mapOutput['categoryIds']]}"
                )
    rptevt.outputs.append(output)

dumper = json_dumper if args.output_format == "json" else yaml_dumper

output_str = dumper.dumps(rptevt)

with open(
    os.path.join("".join(os.path.split(args.excel_file)[0:-1]), rptevtfname + "." + args.output_format),
    "w",
    encoding="utf-8",
) as f:
    f.write(output_str)

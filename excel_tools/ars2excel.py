#!/usr/bin/env python
"""Convert an ARS YAML/JSON file into an Excel workbook matching the ARS template."""

import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

try:
    import yaml
except ImportError:
    yaml = None


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Convert an ARS YAML/JSON file into an Excel workbook using the ARS Template.xlsx structure."
    )
    parser.add_argument(
        "-i",
        "--input_file",
        required=True,
        help="Path to the input ARS file in YAML or JSON format.",
    )
    parser.add_argument(
        "-t",
        "--template_file",
        default=None,
        help="Optional path to the ARS Template.xlsx workbook. If omitted, the script uses the template in the current script directory.",
    )
    return parser.parse_args()


def load_ars_input(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    suffix = path.suffix.lower()
    with path.open("r", encoding="utf-8") as handle:
        if suffix in {".yaml", ".yml"}:
            if yaml is None:
                raise RuntimeError("PyYAML is required to read YAML input files.")
            document = yaml.safe_load(handle)
        elif suffix == ".json":
            document = json.load(handle)
        else:
            raise ValueError("Input file must be .yaml, .yml, or .json")

    if isinstance(document, dict) and "reportingEvent" in document:
        document = document["reportingEvent"]

    if not isinstance(document, dict):
        raise ValueError("ARS input must be a mapping at the root level.")

    return document


def load_template_workbook(template_path: Path):
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    return load_workbook(filename=str(template_path))


def clear_sheet(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def format_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple)):
        return " | ".join(str(item) for item in value if item is not None)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def sheet_headers(ws):
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def expand_headers(ws, column_groups, rows):
    headers = sheet_headers(ws)
    for group in column_groups:
        # group is list of templates like 'groupingId{idx}'
        # find max_idx in rows
        max_idx = 0
        for row in rows:
            for template in group:
                if '{' in template:
                    prefix, suffix = template.split('{idx}', 1)
                    for key in row.keys():
                        if key.startswith(prefix) and key.endswith(suffix):
                            try:
                                idx_str = key[len(prefix):-len(suffix) or None]
                                idx = int(idx_str)
                                max_idx = max(max_idx, idx)
                            except ValueError:
                                pass
                else:
                    if template in row:
                        max_idx = max(max_idx, 1)
        # find current_max in headers
        current_max = 0
        for header in headers:
            for template in group:
                if '{' in template:
                    prefix, suffix = template.split('{idx}', 1)
                    if header.startswith(prefix) and header.endswith(suffix):
                        try:
                            idx_str = header[len(prefix):-len(suffix) or None]
                            idx = int(idx_str)
                            current_max = max(current_max, idx)
                        except ValueError:
                            pass
                else:
                    if header == template:
                        current_max = max(current_max, 1)
        # if max_idx > current_max, insert
        if max_idx > current_max:
            # find position to insert: after the last of the group
            last_col = 0
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(1, col_idx).value
                if header_val in [template.format(idx=current_max) for template in group if '{' in template] + [template for template in group if '{' not in template]:
                    last_col = col_idx
            if last_col:
                insert_pos = last_col + 1
                for idx in range(current_max + 1, max_idx + 1):
                    for template in group:
                        col_name = template.format(idx=idx) if '{' in template else template
                        ws.insert_cols(insert_pos)
                        ws.cell(1, insert_pos).value = col_name
                        insert_pos += 1


def write_rows(ws, rows):
    headers = sheet_headers(ws)
    for row in rows:
        ws.append([format_value(row.get(column)) for column in headers])


def flatten_list_items(list_item, rows, template):
    row = {
        "name": template.get("name"),
        "description": template.get("description"),
        "label": template.get("label"),
        "listItem_level": list_item.get("level"),
        "listItem_name": list_item.get("name"),
        "listItem_description": list_item.get("description"),
        "listItem_label": list_item.get("label"),
        "listItem_order": list_item.get("order"),
        "listItem_analysisId": list_item.get("analysisId"),
        "listItem_outputId": list_item.get("outputId"),
    }
    rows.append(row)
    child_list = list_item.get("sublist")
    if child_list and isinstance(child_list, dict):
        for child in child_list.get("listItems", []):
            flatten_list_items(child, rows, template)


def flatten_categorizations(categorization, parent_id=None, rows=None):
    if rows is None:
        rows = []

    categories = categorization.get("categories") or []
    if categories:
        subcatns = {}
        for category in categories:
            rows.append(
                {
                    "id": categorization.get("id"),
                    "label": categorization.get("label"),
                    "parent_category_id": parent_id,
                    "category_id": category.get("id"),
                    "category_label": category.get("label"),
                }
            )            
            subcatns[category.get("id")] = category.get("subCategorizations") or []
    else:
        rows.append(
            {
                "id": categorization.get("id"),
                "label": categorization.get("label"),
                "parent_category_id": parent_id,
                "category_id": None,
                "category_label": None,
            }
        )

    for cat_id, subcatns in subcatns.items():
        for subcat in subcatns:
            flatten_categorizations(subcat, parent_id=cat_id, rows=rows)

    return rows


def flatten_where_clause(clause, base_row, rows):
    row = dict(base_row)
    row["level"] = clause.get("level")
    row["order"] = clause.get("order")

    condition = clause.get("condition")
    if isinstance(condition, dict):
        row.update(
            {
                "condition_dataset": condition.get("dataset"),
                "condition_variable": condition.get("variable"),
                "condition_comparator": condition.get("comparator"),
                "condition_value": format_value(condition.get("value")),
            }
        )

    compound = clause.get("compoundExpression")
    if isinstance(compound, dict):
        row["compoundExpression_logicalOperator"] = compound.get("logicalOperator")
        if compound.get("whereClauses"):
            rows.append(row)
            for child in compound.get("whereClauses", []):
                if isinstance(child, dict):
                    flatten_where_clause(child, base_row, rows)
            return

    sub_clause_id = clause.get("subClauseId")
    if sub_clause_id is not None:
        row["compoundExpression_subClauseId"] = sub_clause_id

    rows.append(row)


def flatten_group_clause(clause, base_row, rows):
    row = dict(base_row)
    row["group_level"] = clause.get("level")
    row["group_order"] = clause.get("order")

    condition = clause.get("condition")
    if isinstance(condition, dict):
        row.update(
            {
                "group_condition_dataset": condition.get("dataset"),
                "group_condition_variable": condition.get("variable"),
                "group_condition_comparator": condition.get("comparator"),
                "group_condition_value": format_value(condition.get("value")),
            }
        )

    compound = clause.get("compoundExpression")
    if isinstance(compound, dict):
        row["group_compoundExpression_logicalOperator"] = compound.get("logicalOperator")
        if compound.get("whereClauses"):
            rows.append(row)
            for child in compound.get("whereClauses", []):
                if isinstance(child, dict):
                    flatten_group_clause(child, base_row, rows)
            return

    sub_clause_id = clause.get("subClauseId")
    if sub_clause_id is not None:
        row["group_compoundExpression_subClauseId"] = sub_clause_id

    rows.append(row)


def write_reporting_event(ws, event):
    rows = [
        {
            "id": event.get("id"),
            "version": event.get("version"),
            "name": event.get("name"),
            "description": event.get("description"),
            "label": event.get("label"),
        }
    ]
    clear_sheet(ws)
    write_rows(ws, rows)


def write_reference_documents(ws, docs):
    rows = []
    for doc in docs or []:
        rows.append(
            {
                "id": doc.get("id"),
                "name": doc.get("name"),
                "description": doc.get("description"),
                "label": doc.get("label"),
                "location": doc.get("location"),
            }
        )
    clear_sheet(ws)
    write_rows(ws, rows)


def write_terminology_extensions(ws, extensions):
    rows = []
    for ext in extensions or []:
        for sponsor in ext.get("sponsorTerms", []):
            rows.append(
                {
                    "id": ext.get("id"),
                    "enumeration": ext.get("enumeration"),
                    "sponsorTerm_id": sponsor.get("id"),
                    "sponsorTerm_submissionValue": sponsor.get("submissionValue"),
                    "sponsorTerm_description": sponsor.get("description"),
                }
            )
    clear_sheet(ws)
    write_rows(ws, rows)


def write_global_display_sections(ws, sections):
    rows = []
    for section in sections or []:
        for sub in section.get("subSections", []):
            rows.append(
                {
                    "sectionType": section.get("sectionType"),
                    "subSection_id": sub.get("id"),
                    "subSection_text": sub.get("text"),
                }
            )
    clear_sheet(ws)
    write_rows(ws, rows)


def write_categorizations(ws, categories):
    rows = []
    for cat in categories or []:
        rows.extend(flatten_categorizations(cat))
    clear_sheet(ws)
    write_rows(ws, rows)


def write_content_lists(ws, contents):
    rows = []
    for content in contents or []:
        template = {
            "name": content.get("name"),
            "description": content.get("description"),
            "label": content.get("label"),
        }
        for item in (content.get("contentsList", {}).get("listItems") or []):
            flatten_list_items(item, rows, template)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_analysis_sets(ws, analysis_sets):
    rows = []
    for aset in analysis_sets or []:
        base_row = {
            "id": aset.get("id"),
            "name": aset.get("name"),
            "description": aset.get("description"),
            "label": aset.get("label"),
        }
        if aset.get("condition") or aset.get("compoundExpression"):
            clause = {}
            clause.update(aset.get("condition") or {})
            clause["level"] = aset.get("level")
            clause["order"] = aset.get("order")
            if aset.get("compoundExpression"):
                clause["compoundExpression"] = aset.get("compoundExpression")
            if aset.get("condition"):
                clause["condition"] = aset.get("condition")
            flatten_where_clause(clause, base_row, rows)
        elif aset.get("compoundExpression"):
            clause = {"level": aset.get("level"), "order": aset.get("order"), "compoundExpression": aset.get("compoundExpression")}
            flatten_where_clause(clause, base_row, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_analysis_groupings(ws, groupings):
    rows = []
    for grouping in groupings or []:
        base_row = {
            "id": grouping.get("id"),
            "name": grouping.get("name"),
            "description": grouping.get("description"),
            "label": grouping.get("label"),
            "groupingDataset": grouping.get("groupingDataset"),
            "groupingVariable": grouping.get("groupingVariable"),
            "dataDriven": grouping.get("dataDriven"),
        }
        groups = grouping.get("groups", [])
        if not groups:
            rows.append(base_row)
        else:
            for group in groups:
                clause = {
                    "level": group.get("level"),
                    "order": group.get("order"),
                    "condition": group.get("condition"),
                    "compoundExpression": group.get("compoundExpression"),
                    "subClauseId": group.get("subClauseId"),
                }
                row = dict(base_row)
                row["group_id"] = group.get("id")
                row["group_name"] = group.get("name")
                row["group_description"] = group.get("description")
                row["group_label"] = group.get("label")
                flatten_group_clause(clause, row, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_data_subsets(ws, subsets):
    rows = []
    for subset in subsets or []:
        base_row = {
            "id": subset.get("id"),
            "name": subset.get("name"),
            "description": subset.get("description"),
            "label": subset.get("label"),
        }
        clause = {
            "level": subset.get("level"),
            "order": subset.get("order"),
            "condition": subset.get("condition"),
            "compoundExpression": subset.get("compoundExpression"),
            "subClauseId": subset.get("subClauseId"),
        }
        flatten_where_clause(clause, base_row, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def normalize_controlled_term(value):
    if isinstance(value, dict):
        if "controlledTerm" in value:
            return value.get("controlledTerm")
        if "sponsorTermId" in value:
            return value.get("sponsorTermId")
    return value


def write_analyses(ws, analyses):
    rows = []
    for analysis in analyses or []:
        row = {
            "id": analysis.get("id"),
            "version": analysis.get("version"),
            "name": analysis.get("name"),
            "description": analysis.get("description"),
            "label": analysis.get("label"),
            "reason": normalize_controlled_term(analysis.get("reason")),
            "purpose": normalize_controlled_term(analysis.get("purpose")),
            "analysisSetId": analysis.get("analysisSetId"),
            "dataSubsetId": analysis.get("dataSubsetId"),
            "dataset": analysis.get("dataset"),
            "variable": analysis.get("variable"),
            "methodId": analysis.get("methodId"),
            "method_id": analysis.get("methodId"),
            "categoryIds": analysis.get("categoryIds"),
        }
        for idx, grouping in enumerate(analysis.get("orderedGroupings", []), start=1):
            row[f"groupingId{idx}"] = grouping.get("groupingId")
            row[f"resultsByGroup{idx}"] = grouping.get("resultsByGroup")
        for idx, ref in enumerate(analysis.get("referencedAnalysisOperations", []), start=1):
            row[f"referencedAnalysisOperations_referencedOperationId{idx}"] = ref.get("referencedOperationRelationshipId")
            row[f"referencedAnalysisOperations_analysisId{idx}"] = ref.get("analysisId")
            row[f"referencedAnalysisOperations_referencedOperationRelationshipId{idx}"] = ref.get("referencedOperationRelationshipId")
            row[f"referencedAnalysisOperations_operationId{idx}"] = ref.get("operationId")
        rows.append(row)
    column_groups = [
        ['groupingId{idx}', 'resultsByGroup{idx}'],
        ['referencedAnalysisOperations_referencedOperationId{idx}', 'referencedAnalysisOperations_analysisId{idx}', 'referencedAnalysisOperations_referencedOperationRelationshipId{idx}', 'referencedAnalysisOperations_operationId{idx}']
    ]
    expand_headers(ws, column_groups, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_outputs(ws, outputs):
    rows = []
    for output in outputs or []:
        row = {
            "id": output.get("id"),
            "version": output.get("version"),
            "name": output.get("name"),
            "description": output.get("description"),
            "label": output.get("label"),
            "categoryIds": output.get("categoryIds"),
        }
        for idx, disp in enumerate(output.get("displays", []), start=1):
            row[f"display{idx}_id"] = disp.get("display", {}).get("id") if isinstance(disp, dict) else None
        rows.append(row)
    column_groups = [
        ['display{idx}_id']
    ]
    expand_headers(ws, column_groups, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_output_files(ws, outputs):
    rows = []
    for output in outputs or []:
        for file_spec in output.get("fileSpecifications", []):
            rows.append(
                {
                    "output_id": output.get("id"),
                    "name": file_spec.get("name"),
                    "description": file_spec.get("description"),
                    "label": file_spec.get("label"),
                    "location": file_spec.get("location"),
                    "fileType": normalize_controlled_term(file_spec.get("fileType")),
                }
            )
    clear_sheet(ws)
    write_rows(ws, rows)


def page_ref_fields(page_ref):
    if not page_ref or not isinstance(page_ref, dict):
        return {}
    pages = None
    if page_ref.get("refType") == "NamedDestination":
        names = page_ref.get("pageNames")
        if isinstance(names, list):
            pages = "|".join(str(n) for n in names if n is not None)
        else:
            pages = names
    elif page_ref.get("refType") == "PhysicalRef":
        if page_ref.get("pageNumbers") is not None:
            nums = page_ref.get("pageNumbers")
            pages = "|".join(str(n) for n in nums) if isinstance(nums, list) else nums
        elif page_ref.get("firstPage") is not None and page_ref.get("lastPage") is not None:
            pages = f"{page_ref.get('firstPage')}-{page_ref.get('lastPage')}"
    return {
        "pageRef_refType": page_ref.get("refType"),
        "pageRef_label": page_ref.get("label"),
        "pageRef_pages": pages,
    }


def append_document_ref_rows(rows, id_key, id_value, reference_type, doc_ref):
    if not doc_ref or not isinstance(doc_ref, dict):
        return
    base = {
        id_key: id_value,
        "referenceType": reference_type,
        "refDocumentId": doc_ref.get("referenceDocumentId"),
    }
    page_refs = doc_ref.get("pageRefs") or []
    if isinstance(page_refs, dict):
        page_refs = [page_refs]
    if not page_refs:
        rows.append(base)
        return
    for page_ref in page_refs:
        row = dict(base)
        row.update(page_ref_fields(page_ref))
        rows.append(row)


def write_document_refs(ws, refs, id_key):
    rows = []
    for ref in refs or []:
        append_document_ref_rows(rows, id_key, ref.get(id_key), ref.get("referenceType"), ref.get("documentRef"))
    clear_sheet(ws)
    write_rows(ws, rows)


def write_code_parameters(ws, items, id_key, source_key, template=False):
    rows = []
    for item in items or []:
        code_obj = item.get(source_key)
        if not code_obj:
            continue
        for param in code_obj.get("parameters", []):
            row = {
                id_key: item.get("id"),
                "parameter_name": param.get("name"),
                "parameter_description": param.get("description"),
                "parameter_label": param.get("label"),
                "parameter_value": format_value(param.get("value")),
            }
            if template:
                row["parameter_valueSource"] = param.get("valueSource")
            rows.append(row)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_programming_code(ws, items, id_key, source_key, template=False):
    rows = []
    for item in items or []:
        code_obj = item.get(source_key)
        if not code_obj:
            continue
        row = {
            id_key: item.get("id"),
            "context": code_obj.get("context"),
            "specifiedAs": "DocumentRef" if code_obj.get("documentRef") else "Code",
        }
        if template:
            row["templateCode"] = code_obj.get("code")
        else:
            row["code"] = code_obj.get("code")
        rows.append(row)
    clear_sheet(ws)
    write_rows(ws, rows)


def collect_document_refs(document):
    rows = []
    for method in document.get("methods", []):
        for doc in method.get("documentRefs", []):
            rows.append({
                "method_id": method.get("id"),
                "referenceType": "Documentation",
                "refDocumentId": doc.get("referenceDocumentId"),
                "documentRef": doc,
            })
        code_template = method.get("codeTemplate")
        if code_template and code_template.get("documentRef"):
            rows.append({
                "method_id": method.get("id"),
                "referenceType": "ProgrammingCode",
                "refDocumentId": code_template["documentRef"].get("referenceDocumentId"),
                "documentRef": code_template["documentRef"],
            })
    for analysis in document.get("analyses", []):
        for doc in analysis.get("documentRefs", []):
            rows.append({
                "analysis_id": analysis.get("id"),
                "referenceType": "Documentation",
                "refDocumentId": doc.get("referenceDocumentId"),
                "documentRef": doc,
            })
        code_obj = analysis.get("programmingCode")
        if code_obj and code_obj.get("documentRef"):
            rows.append({
                "analysis_id": analysis.get("id"),
                "referenceType": "ProgrammingCode",
                "refDocumentId": code_obj["documentRef"].get("referenceDocumentId"),
                "documentRef": code_obj["documentRef"],
            })
    for output in document.get("outputs", []):
        for doc in output.get("documentRefs", []):
            rows.append({
                "output_id": output.get("id"),
                "referenceType": "Documentation",
                "refDocumentId": doc.get("referenceDocumentId"),
                "documentRef": doc,
            })
        code_obj = output.get("programmingCode")
        if code_obj and code_obj.get("documentRef"):
            rows.append({
                "output_id": output.get("id"),
                "referenceType": "ProgrammingCode",
                "refDocumentId": code_obj["documentRef"].get("referenceDocumentId"),
                "documentRef": code_obj["documentRef"],
            })
    return rows


def write_displays(ws, outputs):
    rows = []
    seen = {}
    for output in outputs or []:
        for ordered in output.get("displays", []):
            display = ordered.get("display")
            if isinstance(display, dict) and display.get("id"):
                seen[display["id"]] = display

    for display in seen.values():
        for section in display.get("displaySections", []):
            for ordered in section.get("orderedSubSections", []):
                row = {
                    "id": display.get("id"),
                    "name": display.get("name"),
                    "description": display.get("description"),
                    "label": display.get("label"),
                    "version": display.get("version"),
                    "displayTitle": display.get("displayTitle"),
                    "displaySection_sectionType": section.get("sectionType"),
                    "displaySection_orderedSubSection_order": ordered.get("order"),
                }
                sub_section = ordered.get("subSection")
                if isinstance(sub_section, dict):
                    row["displaySection_subSection_id"] = sub_section.get("id")
                    row["displaySection_subSection_text"] = sub_section.get("text")
                elif ordered.get("subSectionId"):
                    row["displaySection_subSection_id"] = ordered.get("subSectionId")
                rows.append(row)
    # No iterated columns for displays, so no expand_headers
    clear_sheet(ws)
    write_rows(ws, rows)


def write_analysis_results(ws, analyses, analysis_sets, methods, analysis_groupings):
    analysis_set_labels = {aset.get("id"): aset.get("label") for aset in analysis_sets or []}
    methods_by_id = {method.get("id"): method for method in methods or []}
    analysis_groupings_by_id = {ag.get("id"): ag for ag in analysis_groupings or []}

    rows = []
    for analysis in analyses or []:
        grouping_order = {og['groupingId']: og['order'] for og in analysis.get('orderedGroupings', [])}
        analysis_set_id = analysis.get("analysisSetId")
        method_id = analysis.get("methodId")
        method_label = None
        if method_id and method_id in methods_by_id:
            method_label = methods_by_id[method_id].get("label")

        for result in analysis.get("results", []):
            row = {
                "id": analysis.get("id"),
                "analysisSet_label": analysis_set_labels.get(analysis_set_id),
                "method_id": method_id,
                "method_label": method_label,
                "operation_id": result.get("operationId"),
                "rawValue": result.get("rawValue"),
                "formattedValue": result.get("formattedValue"),
            }

            if method_id and method_id in methods_by_id:
                method = methods_by_id[method_id]
                operation = next((op for op in method.get("operations", []) if op.get("id") == result.get("operationId")), None)
                if operation:
                    row["operation_label"] = operation.get("label")
                    row["operation_resultPattern"] = operation.get("resultPattern")

            for rg in result.get("resultGroups", []):
                grouping_id = rg.get("groupingId")
                if grouping_id in grouping_order:
                    idx = grouping_order[grouping_id]
                    row[f"resultGroup{idx}_groupingId"] = rg.get("groupingId")
                    row[f"resultGroup{idx}_groupId"] = rg.get("groupId")
                    row[f"resultGroup{idx}_groupValue"] = rg.get("groupValue")

                    group_label = None
                    group_id = rg.get("groupId")
                    grouping = analysis_groupings_by_id.get(grouping_id)
                    if grouping and group_id:
                        for group in grouping.get("groups", []):
                            if group.get("id") == group_id:
                                group_label = group.get("label") or group.get("name")
                                break
                    if group_label is not None:
                        row[f"resultGroup{idx}_group_label"] = group_label

            rows.append(row)

    column_groups = [
        ['resultGroup{idx}_groupingId', 'resultGroup{idx}_groupId', 'resultGroup{idx}_groupValue', 'resultGroup{idx}_group_label']
    ]
    expand_headers(ws, column_groups, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_methods(ws, methods):
    rows = []
    for method in methods or []:
        for op in method.get("operations", []):
            row = {
                "id": method.get("id"),
                "name": method.get("name"),
                "description": method.get("description"),
                "label": method.get("label"),
                "operation_id": op.get("id"),
                "operation_name": op.get("name"),
                "operation_description": op.get("description"),
                "operation_label": op.get("label"),
                "operation_order": op.get("order"),
                "operation_resultPattern": op.get("resultPattern"),
            }
            for idx, ref in enumerate(op.get("referencedOperationRelationships", []), start=1):
                row[f"operation_referencedResultRelationships{idx}_id"] = ref.get("id")
                row[f"operation_referencedResultRelationships{idx}_referencedOperationRole"] = normalize_controlled_term(ref.get("referencedOperationRole"))
                row[f"operation_referencedResultRelationships{idx}_operationId"] = ref.get("operationId")
                row[f"operation_referencedResultRelationships{idx}_analysisId"] = ref.get("analysisId")
                row[f"operation_referencedResultRelationships{idx}_description"] = ref.get("description")
                row[f"referencedOperationRelationshipId{idx}"] = ref.get("id")
                row[f"referencedOperationRole{idx}"] = normalize_controlled_term(ref.get("referencedOperationRole"))
                row[f"operationId{idx}"] = ref.get("operationId")
                row[f"referencedOperationRelationshipDescription{idx}"] = ref.get("description")
            rows.append(row)
    column_groups = [
        ['operation_referencedResultRelationships{idx}_id', 'operation_referencedResultRelationships{idx}_referencedOperationRole', 'operation_referencedResultRelationships{idx}_operationId', 'operation_referencedResultRelationships{idx}_analysisId', 'operation_referencedResultRelationships{idx}_description'],
        ['referencedOperationRelationshipId{idx}', 'referencedOperationRole{idx}', 'operationId{idx}', 'referencedOperationRelationshipDescription{idx}']
    ]
    expand_headers(ws, column_groups, rows)
    clear_sheet(ws)
    write_rows(ws, rows)


def write_misc_section(ws, rows):
    clear_sheet(ws)
    write_rows(ws, rows)


def main():
    args = parse_arguments()
    input_path = Path(args.input_file)
    output_path = input_path.with_suffix(".xlsx")
    template_path = Path(args.template_file) if args.template_file else Path(__file__).resolve().parent / "ARS Template.xlsx"

    document = load_ars_input(input_path)
    workbook = load_template_workbook(template_path)

    sections = {
        "ReportingEvent": lambda ws: write_reporting_event(ws, document),
        "ReferenceDocuments": lambda ws: write_reference_documents(ws, document.get("referenceDocuments")),
        "Categorizations": lambda ws: write_categorizations(ws, document.get("analysisOutputCategorizations")),
        "GlobalDisplaySections": lambda ws: write_global_display_sections(ws, document.get("globalDisplaySections")),
        "TerminologyExtensions": lambda ws: write_terminology_extensions(ws, document.get("terminologyExtensions")),
        "MainListOfContents": lambda ws: write_content_lists(ws, [document.get("mainListOfContents")]) if document.get("mainListOfContents") else write_content_lists(ws, []),
        "OtherListsOfContents": lambda ws: write_content_lists(ws, document.get("otherListsOfContents")),
        "AnalysisSets": lambda ws: write_analysis_sets(ws, document.get("analysisSets")),
        "AnalysisGroupings": lambda ws: write_analysis_groupings(ws, document.get("analysisGroupings")),
        "DataSubsets": lambda ws: write_data_subsets(ws, document.get("dataSubsets")),
        "Analyses": lambda ws: write_analyses(ws, document.get("analyses")),
        "Displays": lambda ws: write_displays(ws, document.get("outputs")),
        "AnalysisMethodDocumentRefs": lambda ws: write_document_refs(ws, [r for r in collect_document_refs(document) if "method_id" in r], "method_id"),
        "AnalysisDocumentRefs": lambda ws: write_document_refs(ws, [r for r in collect_document_refs(document) if "analysis_id" in r], "analysis_id"),
        "OutputDocumentRefs": lambda ws: write_document_refs(ws, [r for r in collect_document_refs(document) if "output_id" in r], "output_id"),
        "AnalysisMethodCodeParameters": lambda ws: write_code_parameters(ws, document.get("methods"), "method_id", source_key="codeTemplate", template=True),
        "AnalysisMethodCodeTemplate": lambda ws: write_programming_code(ws, document.get("methods"), "method_id", source_key="codeTemplate", template=True),
        "AnalysisCodeParameters": lambda ws: write_code_parameters(ws, document.get("analyses"), "analysis_id", source_key="programmingCode", template=False),
        "AnalysisProgrammingCode": lambda ws: write_programming_code(ws, document.get("analyses"), "analysis_id", source_key="programmingCode", template=False),
        "OutputCodeParameters": lambda ws: write_code_parameters(ws, document.get("outputs"), "output_id", source_key="programmingCode", template=False),
        "OutputProgrammingCode": lambda ws: write_programming_code(ws, document.get("outputs"), "output_id", source_key="programmingCode", template=False),
        "OutputFiles": lambda ws: write_output_files(ws, document.get("outputs")),
        "Outputs": lambda ws: write_outputs(ws, document.get("outputs")),
        "AnalysisMethods": lambda ws: write_methods(ws, document.get("methods")),
        "AnalysisResults": lambda ws: write_analysis_results(ws, document.get("analyses"), document.get("analysisSets"), document.get("methods"), document.get("analysisGroupings")),
    }

    for sheet_name, writer in sections.items():
        if sheet_name in workbook.sheetnames:
            writer(workbook[sheet_name])

    workbook.save(str(output_path))
    print(f"Generated {output_path}")


if __name__ == "__main__":
    main()
